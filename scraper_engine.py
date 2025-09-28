"""
現実的処理時間対応スクレイピングエンジン
住所取得機能追加版
"""

import time
import random
import logging
import re
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
from urllib.parse import urlparse

# 住所取得対応版のextractorをインポート
from gurunavi_address_extractor import GurunaviAddressExtractor
from gurunavi_label_based_extractor import GurunaviLabelBasedExtractor
from gurunavi_multi_approach_extractor import GurunaviMultiApproachExtractor

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

class ImprovedScraperEngine:
    """段階的動的生成対応スクレイピングエンジンクラス（住所取得対応版）"""
    
    def __init__(self, chrome_manager, prefecture_mapper, config, callback=None):
        self.logger = logging.getLogger(__name__)
        self.chrome_manager = chrome_manager
        self.prefecture_mapper = prefecture_mapper
        self.config = config
        self.callback = callback
        self.driver = None
        self.ua_index = 0
        self.access_count = 0
        self.processed_urls = set()
        
        # Excel保存用の変数
        self.excel_file_path = None
        self.current_results = []
        
        # 並列保存用の設定
        self.save_queue = None
        self.save_executor = None
        self._setup_async_save()
        
        # UA切り替え改善：デフォルト値を30に変更
        if self.config.get('ua_switch_interval', 15) == 15:
            self.config['ua_switch_interval'] = 30
            self.logger.info("UA切り替え間隔を30件に変更しました（60件問題対策）")
        
        # 統計情報
        self.stats = {
            'start_time': None,
            'total_stores': 0,
            'processed_stores': 0,
            'successful_stores': 0,
            'failed_stores': 0,
            'ua_switches': 0,
            'captcha_encounters': 0,
            'ip_restrictions': 0,
            'estimated_completion': None,
            'phone_extraction_failures': 0,
            'address_extraction_failures': 0,  # 住所取得失敗カウント追加
            'success_rate': 1.0
        }
        
        # 時間帯別速度調整
        self.time_multiplier = self._get_time_multiplier()
        
        # プロセス優先度設定
        self._set_process_priority()
    
    def _setup_async_save(self):
        """非同期保存の設定"""
        import queue
        from concurrent.futures import ThreadPoolExecutor
        
        self.save_queue = queue.Queue()
        self.save_executor = ThreadPoolExecutor(max_workers=1)
        
        def save_worker():
            while True:
                item = self.save_queue.get()
                if item is None:
                    break
                try:
                    self._do_save(item['data'], item['path'], item['filename'])
                except Exception as e:
                    self.logger.error(f"非同期保存エラー: {e}")
                finally:
                    self.save_queue.task_done()
        
        self.save_executor.submit(save_worker)
    
    def _set_process_priority(self):
        """プロセス優先度を設定"""
        try:
            import psutil
            import os
            
            p = psutil.Process(os.getpid())
            if os.name == 'nt':  # Windows
                p.nice(psutil.HIGH_PRIORITY_CLASS)
                self.logger.info("プロセス優先度を高に設定しました")
            else:  # Linux/Mac
                p.nice(-5)
                self.logger.info("プロセス優先度を上げました (nice: -5)")
        except Exception as e:
            self.logger.debug(f"プロセス優先度設定失敗: {e}")
    
    def _get_time_multiplier(self):
        """現在時刻に基づく速度調整倍率を取得"""
        try:
            current_hour = datetime.now().hour
            time_config = self.config.get('time_zone_aware', {})
            
            peak = time_config.get('peak_hours', {})
            if peak.get('start', 12) <= current_hour <= peak.get('end', 13):
                return peak.get('multiplier', 1.5)
            
            evening = time_config.get('evening_hours', {})
            if evening.get('start', 18) <= current_hour <= evening.get('end', 20):
                return evening.get('multiplier', 1.3)
            
            safe = time_config.get('safe_hours', {})
            safe_start = safe.get('start', 23)
            safe_end = safe.get('end', 6)
            if current_hour >= safe_start or current_hour <= safe_end:
                return safe.get('multiplier', 0.8)
            
            return 1.0
            
        except Exception as e:
            self.logger.warning(f"時間帯調整取得エラー: {e}")
            return 1.0
    
    def initialize_driver(self):
        """ドライバー初期化（最適化オプション付き）"""
        try:
            user_agent = self.config['user_agents'][self.ua_index]
            
            if hasattr(self.chrome_manager, 'create_optimized_driver'):
                self.driver = self.chrome_manager.create_optimized_driver(
                    headless=True,
                    user_agent=user_agent
                )
            else:
                self.driver = self.chrome_manager.create_driver(
                    headless=True,
                    user_agent=user_agent
                )
            
            self._block_unnecessary_resources()
            
            if self.stats['processed_stores'] < 30:
                self.driver.set_page_load_timeout(20)
            elif self.stats['processed_stores'] < 60:
                self.driver.set_page_load_timeout(25)
            else:
                self.driver.set_page_load_timeout(30)
            
            self.driver.implicitly_wait(8)
            self.driver.set_script_timeout(20)
            
            self.logger.info(f"ドライバー初期化完了 (UA: {self.ua_index}) - 最適化版")
            return True
        except Exception as e:
            self.logger.error(f"ドライバー初期化エラー: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            return False
    
    def _block_unnecessary_resources(self):
        """不要なリソースをブロック"""
        try:
            if self.driver:
                self.driver.execute_cdp_cmd('Network.setBlockedURLs', {
                    'urls': [
                        '*googletagmanager*', 
                        '*google-analytics*',
                        '*doubleclick*', 
                        '*facebook*',
                        '*twitter.com/widgets*',
                        '*platform.twitter*',
                        '*amazon-adsystem*',
                        '*googleapis.com/maps*',
                        '*hotjar*',
                        '*newrelic*',
                        '*clarity.ms*'
                    ]
                })
                self.logger.debug("広告・アナリティクスのブロック設定完了")
        except Exception as e:
            self.logger.debug(f"リソースブロック設定エラー: {e}")
    
    def cleanup(self):
        """クリーンアップ"""
        if self.driver:
            self.chrome_manager.cleanup_driver(self.driver)
            self.driver = None
        
        if self.save_queue:
            self.save_queue.put(None)
        if self.save_executor:
            self.save_executor.shutdown(wait=True)
    
    def switch_user_agent(self):
        """User-Agent切り替え（改善版）"""
        try:
            old_ua = self.ua_index
            self.ua_index = (self.ua_index + 1) % len(self.config['user_agents'])
            self.stats['ua_switches'] += 1
            
            self.logger.info(f"=== UA切り替え開始 (切り替え回数: {self.stats['ua_switches']}) ===")
            
            wait_time = random.uniform(8, 12)
            self.logger.info(f"UA切り替え前の休憩: {wait_time:.1f}秒")
            time.sleep(wait_time)
            
            cookies = None
            try:
                cookies = self.driver.get_cookies()
                self.logger.debug(f"Cookie保存: {len(cookies)}個")
            except:
                pass
            
            self.cleanup()
            if not self.initialize_driver():
                raise Exception("ドライバー再初期化失敗")
            
            self.logger.info("信頼性構築のためトップページアクセス")
            self.driver.get("https://r.gnavi.co.jp")
            time.sleep(random.uniform(3, 5))
            
            if cookies:
                try:
                    for cookie in cookies:
                        if 'expiry' in cookie:
                            del cookie['expiry']
                        self.driver.add_cookie(cookie)
                    self.logger.debug("Cookie復元完了")
                except Exception as e:
                    self.logger.warning(f"Cookie復元失敗: {e}")
            
            additional_wait = random.uniform(5, 8)
            self.logger.info(f"UA切り替え完了: {old_ua} → {self.ua_index}、追加待機: {additional_wait:.1f}秒")
            time.sleep(additional_wait)
            
            if self.stats['processed_stores'] % 60 == 0:
                self.logger.warning("=== 60件処理完了 - 追加の安全対策実行 ===")
                extra_wait = random.uniform(15, 20)
                self.logger.info(f"60件境界での特別待機: {extra_wait:.1f}秒")
                time.sleep(extra_wait)
            
        except Exception as e:
            self.logger.error(f"User-Agent切り替えエラー: {e}")
            raise
    
    def wait_with_cooltime(self):
        """安全なクールタイム待機（動的調整付き）"""
        base_min = self.config['cooltime_min']
        base_max = self.config['cooltime_max']
        
        adjusted_min = base_min * self.time_multiplier
        adjusted_max = base_max * self.time_multiplier
        
        if self.stats['processed_stores'] > 10:
            success_rate = self.stats['successful_stores'] / self.stats['processed_stores']
            if success_rate > 0.9:
                adjusted_min *= 0.8
                adjusted_max *= 0.8
            elif success_rate < 0.5:
                adjusted_min *= 1.2
                adjusted_max *= 1.2
        
        if 58 <= self.stats['processed_stores'] <= 65:
            adjusted_min *= 1.5
            adjusted_max *= 1.5
            self.logger.debug(f"60件境界付近のため待機時間を1.5倍に調整")
        
        cooltime = random.uniform(adjusted_min, adjusted_max)
        
        self.logger.debug(f"クールタイム待機: {cooltime:.1f}秒 (倍率: {self.time_multiplier})")
        time.sleep(cooltime)
    
    def _wait_for_stepwise_content_load(self):
        """段階的コンテンツ読み込み完了待機"""
        try:
            if self._is_list_page():
                self.logger.debug("一覧ページのため段階的読み込みをスキップ")
                return True
            
            self.logger.debug("段階的コンテンツ読み込み待機開始（詳細ページ）")
            
            # スクロールで読み込み誘発
            scroll_positions = [0, 200, 500, 800, 1100, 1400, 800, 400, 0]
            for i, position in enumerate(scroll_positions):
                self.logger.debug(f"スクロール {i+1}/{len(scroll_positions)}: {position}px")
                self.driver.execute_script(f"window.scrollTo(0, {position});")
                time.sleep(0.8)
                
                if i == 4:
                    try:
                        elements_found = len(self.driver.find_elements(By.CSS_SELECTOR, "#info-table, .basic-table"))
                        if elements_found > 0:
                            self.logger.debug("中間チェック: 主要要素検出済み")
                            break
                    except:
                        pass
            
            # 店舗情報エリアにフォーカス
            try:
                info_elements = self.driver.find_elements(By.CSS_SELECTOR, "#info-table, .basic-table, #info-name")
                if info_elements:
                    self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", info_elements[0])
                    time.sleep(1.5)
                    self.logger.debug("店舗情報エリアにフォーカス完了")
            except Exception as e:
                self.logger.debug(f"店舗情報エリアフォーカス失敗: {e}")
            
            time.sleep(1.5)
            
            return True
            
        except Exception as e:
            self.logger.error(f"段階的コンテンツ読み込み待機エラー: {e}")
            return False
    
    def _is_list_page(self):
        """現在のページが一覧ページかどうか判定"""
        try:
            current_url = self.driver.current_url
            return '/rs/' in current_url or '/rs?' in current_url
        except:
            return False
    
    def is_valid_store_url(self, url):
        """店舗URLの有効性チェック"""
        if not url or not isinstance(url, str):
            return False
        
        try:
            parsed = urlparse(url.strip())
            
            if parsed.netloc.lower() != 'r.gnavi.co.jp':
                return False
            
            path = parsed.path.rstrip('/')
            
            invalid_patterns = [
                r'/rs/?$',
                r'/area/',
                r'/city/',
                r'/campaign/',
                r'/lottery/',
                r'/kanjirank',
                r'/mycoupon',
                r'/guide/',
                r'/help/',
                r'/search',
                r'/special/',
                r'/feature/',
                r'/category/',
                r'/genre/',
                r'/apps',
                r'/api/',
                r'/static/',
                r'/css/',
                r'/js/',
                r'/img/',
                r'^/(hokkaido|aomori|iwate|miyagi|akita|yamagata|fukushima|ibaraki|tochigi|gunma|saitama|chiba|tokyo|kanagawa|niigata|toyama|ishikawa|fukui|yamanashi|nagano|gifu|shizuoka|aichi|mie|shiga|kyoto|osaka|hyogo|nara|wakayama|tottori|shimane|okayama|hiroshima|yamaguchi|tokushima|kagawa|ehime|kochi|fukuoka|saga|nagasaki|kumamoto|oita|miyazaki|kagoshima|okinawa)/?$',
            ]
            
            for pattern in invalid_patterns:
                if re.search(pattern, path, re.IGNORECASE):
                    return False
            
            valid_patterns = [
                r'^/[a-zA-Z0-9]{3,20}/?$',
                r'^/[a-zA-Z0-9]{3,20}/(menu|course|map|coupon|photo|plan)/?$'
            ]
            
            for pattern in valid_patterns:
                if re.match(pattern, path, re.IGNORECASE):
                    return True
            
            return False
            
        except Exception as e:
            self.logger.warning(f"URL検証エラー: {url} - {e}")
            return False
    
    def get_base_store_url(self, url):
        """店舗URLのベースURL取得"""
        try:
            parsed = urlparse(url)
            
            if parsed.netloc.lower() != 'r.gnavi.co.jp':
                return None
            
            path_parts = parsed.path.strip('/').split('/')
            
            if len(path_parts) >= 1 and path_parts[0]:
                store_id = path_parts[0]
                if re.match(r'^[a-zA-Z0-9]{3,20}$', store_id):
                    return f"{parsed.scheme}://{parsed.netloc}/{store_id}"
            
            return None
        except Exception:
            return None
    
    def get_store_list(self, prefecture, city, max_count, unlimited):
        """店舗一覧取得"""
        try:
            if not self.initialize_driver():
                raise Exception("ドライバー初期化失敗")
            
            search_url = self.prefecture_mapper.generate_search_url(prefecture, city, page=1)
            self.logger.info(f"検索URL: {search_url}")
            self.logger.info(f"検索エリア: {self.prefecture_mapper.get_area_display_name(prefecture, city)}")
            
            self.driver.get(search_url)
            self._wait_for_list_page_load()
            
            current_url = self.driver.current_url
            page_title = self.driver.title
            self.logger.info(f"ページ読み込み完了 - URL: {current_url}")
            self.logger.info(f"ページタイトル: {page_title}")
            
            if "404" in page_title or "エラー" in page_title or "見つかりません" in page_title:
                raise Exception(f"エラーページが表示されました: {page_title}")
            
            all_store_urls = []
            page_num = 1
            self.processed_urls.clear()
            consecutive_empty_pages = 0
            max_pages = 50 if unlimited else min(20, (max_count // 30) + 5)
            
            while len(all_store_urls) < (float('inf') if unlimited else max_count):
                if self.callback:
                    self.callback({
                        'phase': 'listing',
                        'message': f'ページ {page_num} の店舗URL取得中...',
                        'progress': min((len(all_store_urls) / max_count) * 50, 50) if not unlimited else 0,
                        'current': len(all_store_urls),
                        'target': max_count if not unlimited else '無制限'
                    })
                
                self.logger.info(f"ページ {page_num} の店舗URL取得中...")
                page_store_urls = self._extract_store_urls_from_page()
                
                if not page_store_urls:
                    consecutive_empty_pages += 1
                    self.logger.warning(f"ページ {page_num} で店舗URLが見つかりません（連続{consecutive_empty_pages}回目）")
                    
                    if consecutive_empty_pages >= 3:
                        self.logger.warning("3ページ連続で店舗が見つからないため終了します")
                        break
                else:
                    consecutive_empty_pages = 0
                    
                    new_urls = []
                    for url in page_store_urls:
                        base_url = self.get_base_store_url(url)
                        if base_url not in self.processed_urls:
                            new_urls.append(base_url)
                            self.processed_urls.add(base_url)
                    
                    if not unlimited:
                        remaining = max_count - len(all_store_urls)
                        new_urls = new_urls[:remaining]
                    
                    all_store_urls.extend(new_urls)
                    self.logger.info(f"ページ {page_num}: {len(new_urls)}件取得 (累計: {len(all_store_urls)}件)")
                
                if not unlimited and len(all_store_urls) >= max_count:
                    self.logger.info(f"目標件数に到達しました: {len(all_store_urls)}件")
                    break
                
                if page_num >= max_pages:
                    self.logger.warning(f"最大ページ数({max_pages})に到達しました")
                    break
                
                if page_num % 10 == 0:
                    self._perform_memory_cleanup_light(page_num)
                
                page_num += 1
                next_url = self.prefecture_mapper.generate_search_url(prefecture, city, page=page_num)
                
                self.logger.info(f"次ページへ移動: {next_url}")
                self.driver.get(next_url)
                
                self._wait_for_list_page_load()
                self.wait_with_cooltime()
            
            store_list = []
            for i, url in enumerate(all_store_urls, 1):
                try:
                    parsed = urlparse(url)
                    store_id = parsed.path.strip('/').split('/')[0]
                    name = f"店舗ID: {store_id}"
                except:
                    name = f"店舗 {i}"
                
                store_list.append({
                    'name': name,
                    'url': url
                })
            
            self.logger.info(f"店舗一覧取得完了: {len(store_list)}件")
            return store_list
            
        except Exception as e:
            self.logger.error(f"店舗一覧取得エラー: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            raise
    
    def _wait_for_list_page_load(self):
        """一覧ページ専用の軽量な読み込み待機"""
        try:
            self.logger.debug("一覧ページ読み込み待機開始")
            
            try:
                WebDriverWait(self.driver, 3).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
            except TimeoutException:
                self.logger.warning("一覧ページの基本要素待機タイムアウト")
            
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.5)
            
            links = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='/r.gnavi.co.jp/']")
            
            if links:
                self.logger.debug(f"一覧ページ読み込み完了: {len(links)}個のリンク検出")
                return True
            
            time.sleep(1)
            
            return True
            
        except Exception as e:
            self.logger.warning(f"一覧ページ読み込み待機エラー: {e}")
            return False
    
    def _perform_memory_cleanup_light(self, page_num):
        """軽量版メモリ解放"""
        try:
            self.logger.info(f"=== {page_num}ページ処理完了 - 軽量メモリ解放 ===")
            
            import gc
            
            if len(self.processed_urls) > 500:
                urls_list = list(self.processed_urls)
                self.processed_urls = set(urls_list[-500:])
                self.logger.debug(f"processed_urls制限: {len(urls_list)}件 → 500件")
            
            gc.collect()
            
            self.driver.execute_script("window.localStorage.clear();")
            
            self.logger.info("軽量メモリ解放完了")
            
        except Exception as e:
            self.logger.warning(f"軽量メモリ解放エラー: {e}")
    
    def _extract_store_urls_from_page(self):
        """現在ページから店舗URL抽出"""
        try:
            time.sleep(1)
            
            all_links = self.driver.execute_script("""
                const links = Array.from(document.querySelectorAll('a[href]'));
                return links.map(link => ({
                    href: link.href,
                    text: link.textContent.trim(),
                    title: link.getAttribute('title') || ''
                }));
            """)
            
            store_urls = []
            for link_data in all_links:
                url = link_data['href']
                
                if self.is_valid_store_url(url):
                    normalized_url = url.split('?')[0].rstrip('/')
                    store_urls.append(normalized_url)
            
            unique_urls = list(set(store_urls))
            
            return unique_urls
            
        except Exception as e:
            self.logger.error(f"店舗URL抽出エラー: {e}")
            return []
    
    def get_store_detail(self, url):
        """店舗詳細取得（住所対応版）"""
        try:
            if self.driver is None:
                self.logger.error("Driver is None before creating extractor")
                return self._get_default_detail(url)
            
            if self._quick_error_check():
                self.logger.warning(f"エラーページ検出: {url}")
                return self._get_default_detail(url)
            
            if 58 <= self.stats['processed_stores'] <= 62:
                extra_wait = random.uniform(3, 5)
                self.logger.warning(f"60件境界付近での追加待機: {extra_wait:.1f}秒")
                time.sleep(extra_wait)
            
            success = self._get_with_retry(url)
            if not success:
                return self._get_default_detail(url)
            
            self._wait_for_stepwise_content_load()
            
            # 住所対応版extractorを使用
            from gurunavi_address_extractor import GurunaviAddressExtractor
            extractor = GurunaviAddressExtractor(self.driver, self.logger)
            store_data = extractor.extract_store_data_with_address(url)
            
            if store_data:
                # 統計更新
                if store_data['電話番号'] == '-':
                    self.stats['phone_extraction_failures'] += 1
                    self.logger.warning(f"電話番号取得失敗 (累計: {self.stats['phone_extraction_failures']}件)")
                
                if store_data['住所'] == '-':
                    self.stats['address_extraction_failures'] += 1
                    self.logger.warning(f"住所取得失敗 (累計: {self.stats['address_extraction_failures']}件)")
                
                self.wait_with_cooltime()
                return store_data
            
            # フォールバック
            self.wait_with_cooltime()
            return self._get_default_detail(url)
            
        except Exception as e:
            self.logger.error(f"店舗詳細取得エラー: {e}")
            return self._get_default_detail(url)
    
    def _get_with_retry(self, url, max_retries=2):
        """リトライ機能付きページアクセス"""
        for i in range(max_retries):
            try:
                self.driver.get(url)
                return True
            except TimeoutException:
                self.logger.warning(f"タイムアウト発生 (試行 {i+1}/{max_retries}): {url}")
                if i == 0:
                    try:
                        self.driver.execute_script("window.stop();")
                    except:
                        pass
                    time.sleep(1)
                else:
                    try:
                        self.driver.refresh()
                    except:
                        pass
            except Exception as e:
                self.logger.error(f"ページアクセスエラー: {e}")
                return False
        return False
    
    def _quick_error_check(self):
        """軽量なエラーチェック"""
        try:
            title = self.driver.title
            if any(word in title.lower() for word in ['404', 'error', 'not found', '見つかりません']):
                return True
        except:
            pass
        return False
    
    def _get_default_detail(self, url):
        """デフォルトの店舗データ（住所対応版）"""
        from datetime import datetime
        return {
            'URL': url,
            '店舗名': '取得失敗',
            '電話番号': '-',
            '住所': '-',
            '取得日時': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def _update_estimated_completion(self):
        """完了予想時間の更新"""
        if not self.stats['start_time'] or self.stats['total_stores'] == 0:
            return
        
        elapsed = time.time() - self.stats['start_time']
        if self.stats['processed_stores'] > 0:
            avg_time_per_store = elapsed / self.stats['processed_stores']
            remaining_stores = self.stats['total_stores'] - self.stats['processed_stores']
            estimated_remaining = remaining_stores * avg_time_per_store
            
            self.stats['estimated_completion'] = datetime.now() + timedelta(seconds=estimated_remaining)
    
    def get_processing_stats(self):
        """処理統計情報取得（住所対応版）"""
        elapsed = 0
        if self.stats['start_time']:
            elapsed = time.time() - self.stats['start_time']
        
        return {
            '経過時間': f"{elapsed/60:.1f}分",
            '処理済み店舗数': self.stats['processed_stores'],
            '成功店舗数': self.stats['successful_stores'],
            '失敗店舗数': self.stats['failed_stores'],
            '電話番号取得失敗': self.stats['phone_extraction_failures'],
            '住所取得失敗': self.stats['address_extraction_failures'],
            'UA切り替え回数': self.stats['ua_switches'],
            'CAPTCHA遭遇回数': self.stats['captcha_encounters'],
            'IP制限遭遇回数': self.stats['ip_restrictions'],
            '平均処理時間/店舗': f"{elapsed/max(self.stats['processed_stores'], 1):.1f}秒",
            '完了予想時刻': self.stats['estimated_completion'].strftime('%H:%M:%S') if self.stats['estimated_completion'] else 'N/A',
            '現在時間帯倍率': f"{self.time_multiplier}x"
        }
    
    def start_processing(self, store_list, search_params):
        """メイン処理開始（住所対応版）"""
        self.stats['start_time'] = time.time()
        self.stats['total_stores'] = len(store_list)
        self.current_results = []
        
        self.logger.info(f"=== 処理開始 (住所取得対応版) ===")
        self.logger.info(f"対象店舗数: {len(store_list)}")
        self.logger.info(f"時間帯倍率: {self.time_multiplier}x")
        self.logger.info(f"UA切り替え間隔: {self.config['ua_switch_interval']}件")
        self.logger.info(f"予想処理時間: {len(store_list) * 8 * self.time_multiplier / 60:.1f}分")
        
        save_dir = Path(search_params['save_path'])
        save_dir.mkdir(parents=True, exist_ok=True)
        filename = search_params['filename']
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        self.excel_file_path = save_dir / filename
        
        # 住所カラムを含めてExcel初期化
        self._initialize_excel_with_urls(store_list)
        
        self._init_memory_monitoring()
        
        if not self.initialize_driver():
            raise Exception("ドライバー初期化失敗")
        
        try:
            for idx, store in enumerate(store_list, 1):
                if self.callback:
                    progress_data = {
                        'phase': 'detail',
                        'message': f'店舗詳細取得中 ({idx}/{len(store_list)}): {store["name"]}',
                        'progress': (idx / len(store_list)) * 100,
                        'current': idx,
                        'total': len(store_list),
                        'stats': self.get_processing_stats()
                    }
                    self.callback(progress_data)
                
                detail = self.get_store_detail(store['url'])
                
                self._update_excel_row(idx, detail)
                
                self.stats['processed_stores'] = idx
                if detail['店舗名'] != '取得失敗' and detail['店舗名'] != '-':
                    self.stats['successful_stores'] += 1
                else:
                    self.stats['failed_stores'] += 1
                
                self.stats['success_rate'] = self.stats['successful_stores'] / self.stats['processed_stores']
                
                self._update_estimated_completion()
                
                if idx % 50 == 0:
                    self._check_memory_usage()
                
                ua_interval = self.config.get('ua_switch_interval', 30)
                
                if idx == 60:
                    self.logger.warning("=== 60件処理完了 - 特別なUA切り替え実行 ===")
                    try:
                        self.switch_user_agent()
                    except Exception as e:
                        self.logger.error(f"60件境界でのUA切り替えエラー: {e}")
                        time.sleep(10)
                
                elif idx % ua_interval == 0 and idx < len(store_list) and idx != 60:
                    self.logger.info(f"定期UA切り替え実行 ({idx}件処理完了)")
                    try:
                        self.switch_user_agent()
                    except Exception as e:
                        self.logger.error(f"UA切り替えエラー: {e}")
                        break
            
            final_stats = self.get_processing_stats()
            self.logger.info("=== 処理完了統計 ===")
            for key, value in final_stats.items():
                self.logger.info(f"{key}: {value}")
            
            self._save_stats_to_excel()
            
            df = pd.read_excel(self.excel_file_path, sheet_name='店舗詳細')
            self.current_results = df.to_dict('records')
            
            return self.current_results
            
        finally:
            self.cleanup()
    
    def _initialize_excel_with_urls(self, store_list):
        """ExcelファイルをURL一覧で初期化（住所対応版）"""
        try:
            self.logger.info(f"ExcelファイルをURL一覧で初期化: {self.excel_file_path}")
            
            # 住所カラムを追加
            data = []
            for store in store_list:
                data.append({
                    'URL': store['url'],
                    '店舗名': '',
                    '電話番号': '',
                    '住所': '',  # 追加
                    '取得日時': ''
                })
            
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(self.excel_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='店舗詳細', index=False)
                
                worksheet = writer.sheets['店舗詳細']
                worksheet.column_dimensions['A'].width = 60  # URL
                worksheet.column_dimensions['B'].width = 30  # 店舗名
                worksheet.column_dimensions['C'].width = 15  # 電話番号
                worksheet.column_dimensions['D'].width = 40  # 住所（新規追加）
                worksheet.column_dimensions['E'].width = 20  # 取得日時
            
            self.logger.info(f"URL一覧で初期化完了: {len(store_list)}件")
            
        except Exception as e:
            self.logger.error(f"Excel初期化エラー: {e}")
    
    def _update_excel_row(self, row_number, detail):
        """Excelの特定行を更新（住所対応版）"""
        try:
            from openpyxl import load_workbook
            
            book = load_workbook(self.excel_file_path)
            sheet = book['店舗詳細']
            
            excel_row = row_number + 1
            
            # データを更新（住所カラムを追加）
            sheet.cell(row=excel_row, column=2, value=detail.get('店舗名', ''))      # B列: 店舗名
            sheet.cell(row=excel_row, column=3, value=detail.get('電話番号', ''))    # C列: 電話番号
            sheet.cell(row=excel_row, column=4, value=detail.get('住所', ''))        # D列: 住所（新規）
            sheet.cell(row=excel_row, column=5, value=detail.get('取得日時', ''))    # E列: 取得日時
            
            book.save(self.excel_file_path)
            book.close()
            
            if row_number % 10 == 0:
                status = "成功" if detail.get('店舗名') not in ['', '-', '取得失敗'] else "失敗"
                self.logger.info(f"Excel更新: {row_number}行目 - {status}")
            
        except Exception as e:
            self.logger.error(f"Excel行更新エラー (行{row_number}): {e}")
    
    def _save_stats_to_excel(self):
        """処理統計を同じExcelの「処理統計」シートに保存"""
        try:
            from openpyxl import load_workbook
            
            stats = self.get_processing_stats()
            stats_data = []
            for key, value in stats.items():
                stats_data.append({'項目': key, '値': str(value)})
            
            df_stats = pd.DataFrame(stats_data)
            
            if self.excel_file_path.exists():
                book = load_workbook(self.excel_file_path)
                
                if '処理統計' in book.sheetnames:
                    del book['処理統計']
                
                book.save(self.excel_file_path)
                book.close()
                
                with pd.ExcelWriter(self.excel_file_path, engine='openpyxl', mode='a') as writer:
                    df_stats.to_excel(writer, sheet_name='処理統計', index=False)
                    
                    worksheet = writer.sheets['処理統計']
                    worksheet.column_dimensions['A'].width = 25
                    worksheet.column_dimensions['B'].width = 30
            
            self.logger.info("処理統計を保存しました")
            
        except Exception as e:
            self.logger.warning(f"処理統計保存エラー: {e}")
    
    def _init_memory_monitoring(self):
        """メモリ監視の初期化"""
        try:
            import psutil
            import gc
            
            gc.collect()
            gc.set_threshold(700, 10, 10)
            
            self.process = psutil.Process()
            initial_memory = self.process.memory_info().rss / 1024 / 1024
            self.logger.info(f"初期メモリ使用量: {initial_memory:.1f}MB")
        except Exception as e:
            self.logger.debug(f"メモリ監視初期化エラー: {e}")
    
    def _check_memory_usage(self):
        """メモリ使用量チェック"""
        try:
            import psutil
            import gc
            
            process_memory = self.process.memory_percent()
            system_memory = psutil.virtual_memory().percent
            
            self.logger.debug(f"メモリ使用率 - プロセス: {process_memory:.1f}%, システム: {system_memory:.1f}%")
            
            if process_memory > 5.0:
                gc.collect()
                self.logger.info("ガベージコレクション実行")
            
            if system_memory > 80:
                self.logger.warning(f"システムメモリ使用率が高い: {system_memory:.1f}%")
                if system_memory > 85 and self.stats['processed_stores'] % 30 == 0:
                    self.logger.warning("メモリ逼迫のためドライバー再起動を推奨")
        
        except Exception as e:
            self.logger.debug(f"メモリチェックエラー: {e}")
    
    def save_results(self, results, save_path, filename):
        """最終結果をExcelに保存（住所対応版）"""
        try:
            save_dir = Path(save_path)
            save_dir.mkdir(parents=True, exist_ok=True)
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            full_path = save_dir / filename
            
            if full_path.exists():
                from openpyxl import load_workbook
                book = load_workbook(full_path)
                
                with pd.ExcelWriter(full_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    writer.book = book
                    
                    df = pd.DataFrame(results)
                    df.to_excel(writer, sheet_name='店舗詳細', index=False)
                    
                    if hasattr(self, 'get_processing_stats'):
                        stats_df = pd.DataFrame([self.get_processing_stats()]).T
                        stats_df.columns = ['値']
                        stats_df.to_excel(writer, sheet_name='処理統計')
                    
                    for sheet_name in ['店舗詳細', '処理統計']:
                        if sheet_name in writer.sheets:
                            worksheet = writer.sheets[sheet_name]
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                worksheet.column_dimensions[column_letter].width = adjusted_width
            else:
                with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                    df = pd.DataFrame(results)
                    df.to_excel(writer, sheet_name='店舗詳細', index=False)
                    
                    if hasattr(self, 'get_processing_stats'):
                        stats_df = pd.DataFrame([self.get_processing_stats()]).T
                        stats_df.columns = ['値']
                        stats_df.to_excel(writer, sheet_name='処理統計')
            
            self.logger.info(f"最終結果保存: {full_path}")
            
        except Exception as e:
            self.logger.error(f"結果保存エラー: {e}")
            raise